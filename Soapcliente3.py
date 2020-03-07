# -*- coding: UTF-8 -*-
from suds.client import Client
import sqlite3 as lite
import os
import sys
import time
from openpyxl import load_workbook
from imp import reload
reload(sys)  # Reload does the trick!
sys.setdefaultencoding('UTF-8')
cl = Client('http://192.168.1.34:7789/?wsdl')
cl.options.cache.clear()

# res = cl.service.ExecCommand("DELETE FROM ubicaciones")
# if (res == 0):
#     print ("borrado correcto")
# else:
#     print("Operacion incorrecta")

# wb = load_workbook(filename = '//Denebola/share/ultimos importados/almv.xlsx')
# print(wb.get_sheet_names())  #muestra todas las hojas del fichero
# ws = wb.active
# ws = wb.get_sheet_by_name('datos') #abre la hoja apuntada

# print('-------------------------')
# fila = []
# Filas_ = []
# #all_rows = ws.rows.count
# nfilas = ws.max_row 
# ncol = ws.max_column + 1
# print(" Numero de lineas en el fichero " + str(nfilas))
# print(" Numero de columnas en el fichero " + str(ncol))
# ind = 0
# for i in range(1,nfilas ):
#     for j in range(ncol):
       
#         if i > 1 and j > 0:
            
#             fila.insert(ind , ws.cell(row = i ,column = j).value)
#             ind += 1
            
         

#     if i > 1 and j > 0:
#         Filas_.append(fila)
#         fila = []

# for j in Filas_:
#     #time.sleep(0.1)
    

#     sql = "INSERT INTO ubicaciones(modulo,estanteria,ubicacion,division,codigo,nombre,cod_barras,pacto,stock,minimo,dc) VALUES('"+j[0]+"','"+j[1]+"','"+j[2]+"','"+j[3]+"','"+j[4]+"','"+j[5]+"','"+j[6]+"','"+j[7]+"','"+j[8]+"','"+j[9]+"','"+j[10]+"')"
# #cl.service.ExecCommand('INSERT INTO ubicaciones VALUES (?,?,?,?,?,?,?,?,?,?,?),'+ j)
# #valor = cl.service.ExecCommand("INSERT INTO ubicaciones(modulo,estanteria,ubicacion,division,codigo,nombre,cod_barras,pacto,stock,minimo,dc) VALUES('"+j[0]+"','"+j[1]+"','"+j[2]+"','"+j[3]+"','"+j[4]+"','"+j[5]+"','"+j[6]+"','"+j[7]+"','"+j[8]+"','"+j[9]+"','"+j[10]+"')")

#     valor = cl.service.ExecCommand(sql)
# if(valor == 0):
#    print("EXITO")
# else:
#    print("FALLO")
##for j in Filas_:
##    try:
##        cl.service.ExecCommand('INSERT INTO ubicaciones VALUES (?,?,?,?,?,?,?,?,?,?,?)',j)
##
##    except :
##        print("Fallo al insertar datos en el servidor")
##        
    
    




wb = load_workbook(filename = 'C:/tmp/150419.xlsx')
print(wb.get_sheet_names())
ws = wb.get_sheet_by_name('Master') #abre la hoja apuntada
wl = None
print('-------------------------')
user = []
Usuarios = []

nfilas = ws.max_row 
ncol = ws.max_column 
print(" Numero de lineas en el fichero " + str(nfilas))
print(" Numero de columnas en el fichero " + str(ncol))
ct = 0
for i in range(nfilas):
	if(wl != None and len( user ) > 0):
		Usuarios.insert(ct , user)
		ct += 1
		del user
		user=[]
		
	del wl
	for j in range(ncol):#ncol
		wl = ws.cell(row=i + 2, column=j + 1)
		user.append(wl.value)
	#print(Usuarios[i])


		
#res = cl.service.ExecCommand("delete from articulos");	
#print(res)
#res = cl.service.ExecCommand("DELETE FROM sqlite_sequence WHERE name = 'articulos'");	
#print(res)

print("Numero Articulos: " + str(len(Usuarios)))
print('-------------------------')
#sheet_ranges = wb['Users']
#print(sheet_ranges['A1'].value)

print("Insertando filas espera....")
#db = lite.connect('/home/peli/SqliteDbs/SqliteData.s3db')
#cursor = db.cursor()
contai = 0
contador = 0
for i in Usuarios:
	contai += 1;
	contador += 1
	#time.sleep(2)
	#print(i[0])
	#print(i[1])
	#print(i[2])
	#print(i[3])
	#print(i[4])
	res = cl.service.ExecCommand("INSERT INTO articulos(nombre , codigo , cod_barras ,fk_familia , imputacion )VALUES('"+str(i[1].encode('UTF-8'))+"','"+str(i[0])+"','"+str(i[0])+"',4 , 0)")
	#print("Retorno = " , res)
time.sleep(1)
print("Operacion finalizada.")



#res = cl.service.ExecCommand("INSERT INTO articulos(nombre , codigo , cod_barras ,fk_familia , imputacion )VALUES('"+str(i[0].encode('cp1252'))+"','"+str(i[1])+"','"+str(i[2])+"','"+str(i[3])+"','"+str(i[4])+"')")


