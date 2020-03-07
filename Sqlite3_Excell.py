# -*- coding: utf-8 -*-
import sqlite3 as lite
import os
import sys
#import openpyxl #leer excell
from openpyxl import load_workbook

reload(sys)  # Reload does the trick!
sys.setdefaultencoding('utf-8')

wb = load_workbook(filename = '/home/peli/SqliteDbs/usuarios.xlsx')
print(wb.get_sheet_names())  #muestra todas las hojas del fichero
ws = wb.get_sheet_by_name('Users') #abre la hoja apuntada

wl = ws.cell('A2')
print(wl.value)
wl = ws.cell('A22') 
wl = None
#print(wl.value)
print('-------------------------')
user = []
Usuarios = []
#all_rows = ws.rows.count
nfilas = ws.max_row - 1
ncol = ws.max_column 
print(" Numero de lineas en el fichero " + str(nfilas))
print(" Numero de columnas en el fichero " + str(ncol))
for i in range(nfilas):
        if(wl != None):
		Usuarios.append(user)
		#print(Usuarios[i])

	del wl
	for j in range(ncol):
		wl = ws.cell(row=i +1 ,column=j +1)
		user.append(wl.value)
		#print(user.value)
	
    
print('-------------------------')
#sheet_ranges = wb['Users']
#print(sheet_ranges['A1'].value)


db = lite.connect('/home/peli/SqliteDbs/data.s3db')
cursor = db.cursor()
nombre = ""
login = ""
rfid = ""
grupo = ""
x = 0
for i in Usuarios:
	#print(i.value)
	#print()
	for j in i:
                #print(j )

		#cursor.execute("INSERT INTO usuarios(nombre , login , rfid , grupo)VALUES("'+j[0]+'","'+j[1]+'","'+j[2]+'","'+j[3]')")
		#cursor.execute("INSERT INTO usuarios)VALUES(?,?,?,?)",j[0],j[1],j[2],j[3])
#cursor.commit()
#for registro in cursor:
#print(registro)
             
