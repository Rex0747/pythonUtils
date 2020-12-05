# -*- coding: utf-8 -*-
"""
Created on Tue Dec  1 20:57:19 2020

@author: Pedro
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter


class Excell:
    # openpyxl
    nombre = ''

    def __init__(self, nombre_):
        self.lista = []
        global nombre
        nombre = nombre_
        print('NombreFichero: ' + nombre )
        try:
            self.wb = openpyxl.load_workbook(nombre)  # workbook()
            hoja = self.mostrar_sheets()  #imprime hojas
            self.ws = self.wb[ hoja[0] ]

            #print('se abrio fichero excell '+nombre +' correctamente.')
        except:
            self.wb = openpyxl.Workbook()  #cambiado de workbook()
            #self.ws = self.wb.create_sheet()
            hoja = self.mostrar_sheets()  #imprime hojas
            
            self.ws = self.wb[ hoja[0] ]
            self.ws.title = hoja[0]
            
            self.salvarexcell() #self.salvarexcell2()
            #print('se ha creado el fichero excell '+nombre+' correctamente.')
        # global ws
        self.ws = self.wb.active

    def createsheet(self, titulo):

        hoja = self.wb.create_sheet()
        hoja.title = titulo
        # hoja.sheet_properties.tabcolor = "aaaaaa"
        #self.wb.save(nombre)

    def deleteSheet(self, nombre):
        #hoja= self.wb.get_sheet_by_name(nombre)
        del self.wb[nombre]


    def cambiar_hoja(self, nombre):
        self.ws = self.wb[nombre]

    def mostrar_sheets(self):
        # print( wb.get_sheet_names( ) )
        h = self.wb.sheetnames
        print(h)
        return h

    def mostrar_celda(self, fila, columna):
        return self.ws.cell(row=fila, column=columna).value

    def ver_celda(self, celda):
        return self.ws[celda].value

    def escribir_celda(self, celda, val):
        # self.ws = self.wb[ 'tr' ]     #get_sheet_by_name("prb1")
        self.ws[celda].value = val
        #self.wb.save(nombre)

    def leer_rango(self, rfila, rcolumna):
        ret = []
        lista = self.ws[rfila: rcolumna]
        for i in lista:
            for j in i:
                ret.append(j.value)
        return ret

    def leer_fichero(self):
        ret = []
        item = []
        lista = self.ws.rows
        for i in lista:
            for j in i:
                item.append( j.value )
            ret.append( item )
            item = None
            item = []
        ret.pop( 0 )
        return ret

    def leer_fichero2(self):
        #lista = self.ws.rows
        #return lista
        cmps = None
        self.lista = []
        mtx = self.ws.rows
        try:
            for row in mtx:
                if row != '':
                    m = row[0].value
                    e = row[1].value
                    u = row[2].value
                    d = row[3].value
                    codigo = row[4].value
                    nombre = row[5].value
                    pacto = row[6].value
                    minimo = row[7].value
                    dc = row[8].value
                    gfh = row[9].value
                    disp = row[10].value
                    hosp = row[11].value
                    cmps = campos(m,e,u,d,codigo,nombre,pacto,minimo,dc,gfh,disp,hosp)
                    #cmps = (m+'#'+e+'#'+u+'#'+d+'#'+codigo+'#'+nombre+'#'+pacto+'#'+minimo+'#'+dc+'#'+gfh+'#'+disp+'#'+hosp).split('#')
                    self.lista.append(cmps)
        except Exception as e:
            return []

        return self.lista[1 : ]

    def getNombreColumna( nCol): #consigue la letra
        return get_column_letter(nCol)
        

    def leer_fila(self):
        ret = []
        lista = self.ws.rows
        for i in lista:
            for j in i:
                ret.append(j.value)
        return ret


    def leer_columna(self):
        ret = []
        lista = self.ws.columns
        for i in lista:
            for j in i:
                ret.append(j.value)
        return ret

    def merge(self, rfila, rcolumna):
        rango = rfila + ':' + rcolumna
        self.ws.merge_cells(rango)
        

    def unmerge(self, rfila, rcolumna):
        rango = rfila + ':' + rcolumna
        self.ws.unmerge_cells(rango)
        

    def insertar_imagen(self, ruta, celda):
        img = openpyxl.drawing.image.image(ruta)
        self.ws.add_image(img, celda)
        

    def insertar_rangofila(self, rango, fila, columna):
        for i, value in enumerate(rango):
            self.ws.cell(column=columna + i, row=fila, value=value)
            #print('Fila: '+ str(fila) + '  Columna: '+ str(columna) + '  Value: '+ str(value))

    def insertar_rangocolumna(self, rango, fila, columna):
        for i, value in enumerate(rango):
            self.ws.cell(column=columna, row=fila + i, value=value)
        

    def cambiarcolorfila(self, fila, col, inicol, fincol):
        for i in range(col):
            self.ws.cell(row=fila, column=i + 1).fill = patternfill(start_color=inicol, end_color=fincol,
                                                                    fill_type='solid')
        

    def cambiarcolorcolumna(self, fila, col, inicol, fincol):
        for i in range(fila):
            self.ws.cell(row=i + 1, column=col).fill = patternfill(start_color=inicol, end_color=fincol, fill_type='solid')
        

    def getnumerofilas(self):
        return self.ws.max_row

    def getnumerocolumnas(self):
        return self.ws.max_column

    def salvarexcell(self):
        self.wb.save( nombre )

    def salvarexcell2(self):
        self.wb.save( MEDIA_ROOT +'/' + nombre +'.xlsx' )
        
    def salvarexcell3(self):
        self.wb.save( MEDIA_ROOT +'/' + nombre +'.xlsx' )

    def __repr__(self):
        return str(self.lista)

    def __str__(self):
        return str(self.lista)