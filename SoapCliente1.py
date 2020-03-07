
# -*- coding: UTF-8 -*-
from suds.client import Client
import sqlite3 as lite
import os
import sys
import time
from openpyxl import load_workbook
from imp import reload
reload(sys)  # Reload does the trick!
#sys.setdefaultencoding('UTF-8')
cl = Client('http://192.168.1.221:7789/?wsdl')
cl.options.cache.clear()
wb = load_workbook(filename = 'C:/share/MasterArticulosHupa.xlsx')
wb.get_sheet_names()
ws = wb.get_sheet_by_name('ModSqlite2') #abre la hoja apuntada
wl = None
print('-------------------------')
user = []
Usuarios = []

nfilas = ws.max_row 
ncol = ws.max_column 
print(" Numero de lineas en el fichero " + str(nfilas))
print(" Numero de columnas en el fichero " + str(ncol))
ct = 0
for i in range(nfilas):
	if(wl != None):
		Usuarios.insert(ct , user)
		ct += 1
		del user
		user=[]
		#print(Usuarios[i])
	del wl
	for j in range(ncol):#ncol
		wl = ws.cell(row=i + 2, column=j + 1)
		user.append(wl.value)
		#print(user.value)
	
print(len(Usuarios)) 
print('-------------------------')
#sheet_ranges = wb['Users']
#print(sheet_ranges['A1'].value)

print("Insertando filas espera....")
#db = lite.connect('/home/peli/SqliteDbs/SqliteData.s3db')
#cursor = db.cursor()
contai = 0
contador = 0
for i in Usuarios:
	contai += 1;
	contador += 1
	#time.sleep(2)
	#print(i[0])
	#print(i[1])
	#print(i[2])
	#print(i[3])
	#print(i[4])
	res = cl.service.ExecCommand("INSERT INTO articulos(nombre , codigo , cod_barras ,fk_familia , imputacion )VALUES('"+str(i[0].encode('UTF-8'))+"','"+str(i[1])+"','"+str(i[2])+"','"+str(i[3])+"','"+str(i[4])+"')")
	print("Retorno = " , res)
time.sleep(1)
print("Operacion finalizada.")



#res = cl.service.ExecCommand("INSERT INTO articulos(nombre , codigo , cod_barras ,fk_familia , imputacion )VALUES('"+str(i[0].encode('cp1252'))+"','"+str(i[1])+"','"+str(i[2])+"','"+str(i[3])+"','"+str(i[4])+"')")


