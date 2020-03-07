# -*- coding: utf-8 -*-
#import openpyxl #leer excell
from openpyxl import load_workbook
import sqlite3 as lite
import MySQLdb
import sys

cnx = MySQLdb.connect(host='192.168.1.34', user='peli', passwd='perikillo', db='almacenes')
                              
cur = cnx.cursor()

# Usa todas las sentencias SQL que quieras
cur.execute("SELECT * FROM usuarios")

# Imprimir la primera columna de todos los registros
for row in cur.fetchall():
    print row[2]                            
cnx.close()


def iterrows(row):
    #for row in ws.iter_rows():
    yield [cell.value for cell in row]

print sys.stdout.encoding
#sys.setdefaultencoding('utf8')

wb = load_workbook(filename = 'C:\Users\peli\Downloads\CatalogoEstrellasDobles.xlsx')
print(wb.get_sheet_names())  #muestra todas las hojas del fichero
ws = wb.active
ws = wb.get_sheet_by_name('Catalogo') #abre la hoja apuntada

#wl = ws.cell('A1')
#print(wl.value)
#wl = ws.cell('B1')
#print(wl.value)
#wl = None
#print(wl.value)
print('-------------------------')
fila = []
Filas_ = []
#all_rows = ws.rows.count
nfilas = ws.max_row
ncol = ws.max_column + 1
print(" Numero de lineas en el fichero " + str(nfilas))
print(" Numero de columnas en el fichero " + str(ncol))
ind = 0
for i in range(1,nfilas ):
    for j in range(ncol):

        if i > 1 and j > 0:

            fila.insert(ind , ws.cell(row = i ,column = j).value )
            ind += 1



    if i > 1 and j > 0:
        Filas_.append(fila)
        fila = []
print(Filas_)

# db = lite.connect('C:/Users/peli/Documents/SqliteDbs/Modulos.s3db')
# cursor = db.cursor()
#
# #reg = (1,1,1,1,'09878','aspirinas',"9385738945PO9888",120,100,12,1)
#
#
# for j in Filas_:
#     print(type(j))
#     #print(j[0],j[1],j[2],j[3],j[4],j[5],j[6],j[7],j[8],j[9],j[10])
#     cursor.execute("INSERT INTO ubicaciones(modulo,estanteria,ubicacion,division,codigo,nombre,cod_barras,pacto,stock,minimo,dc) VALUES('"+j[0]+"','"+j[1]+"','"+j[2]+"','"+j[3]+"','"+j[4]+"','"+j[5]+"','"+j[6]+"','"+j[7]+"','"+j[8]+"','"+j[9]+"','"+j[10]+"')")
#     #cursor.execute('INSERT INTO ubicaciones VALUES (?,?,?,?,?,?,?,?,?,?,?)', j) #[0],j[1],j[2],j[3],str(j[4]),str(j[5]),str(j[6]),float(j[7]),float(j[8]),float(j[9]),j[10])
#
#
#     #for registro in cursor:
#     #	print(registro)
#     #print("\n")
#
# db.commit()
# db.close()
