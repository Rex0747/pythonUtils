# -*- coding: utf-8 -*-
"""
Created on Tue Jan 26 18:49:53 2021

@author: Pedro
"""

import openpyxl
from openpyxl.utils import get_column_letter

libro = openpyxl.load_workbook('C:/temp/PactosHupa/H2CC.xlsx')
hoja = libro.worksheets[0]
print(get_column_letter(3))

from collections import defaultdict

FDE = [1,1,2,2,3,4,4,3,3,5]

aux = defaultdict(list)
for index, item in enumerate(FDE):
    aux[item].append(index)
result = {item: indexs for item, indexs in aux.items() if len(indexs) > 2 or len(indexs)==1}
print(result)
