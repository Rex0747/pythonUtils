from openpyxl import * #Workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

class Excell:
    
    #openpyxl
    nombre = ''
    
    def __init__( self  , nombre_ , hoja ):
        
        global nombre
        nombre = nombre_
        try:
            self.wb = openpyxl.load_workbook( nombre )     #workbook()
            self.ws = self.wb[ hoja ]

            print("se abrio fichero correctamente.")
        except:
            self.wb =openpyxl.workbook()
            self.ws = self.wb.create_sheet()
            #self.ws.title( hoja )
            print("se ha creado el fichero correctamente.")
        #global ws
        self.ws = self.wb.active


    def createsheet(self , titulo):

        hoja = self.wb.create_sheet()
        hoja.title = titulo
        #hoja.sheet_properties.tabcolor = "aaaaaa"
        self.wb.save( nombre )

    def cambiar_hoja(self , nombre ):
        self.ws = self.wb[ nombre ]

    def mostrar_sheets(self):
        #print( wb.get_sheet_names( ) )
        h = self.wb.sheetnames
        print( h )

    def mostrar_celda( self , fila , columna):
        return self.ws.cell( row = fila , column = columna ).value

    def ver_celda( self , celda ):
        return self.ws[ celda ].value

    def escribir_celda(self , celda  , val):
        #self.ws = self.wb[ 'tristitia' ]     #get_sheet_by_name("prb1")
        self.ws[ celda ].value = val
        self.wb.save( nombre )

    def leer_rango(self , rfila , rcolumna):
        ret = []
        lista = self.ws[ rfila : rcolumna  ]
        for i in lista:
            for j in i:
                ret.append( j.value )
        return ret

    def leer_fila( self ):
        ret = []
        lista = self.ws.rows
        for i in lista:
            for j in i:
                ret.append( j.value )
        return ret

    def leer_columna( self ):
        ret = []
        lista = self.ws.columns
        for i in lista:
            for j in i:
                ret.append( j.value )
        return ret

    def merge( self , rfila , rcolumna ):
        rango = rfila + ':' + rcolumna
        self.ws.merge_cells( rango )
        self.wb.save( nombre )

    def unmerge( self , rfila , rcolumna ):
        rango = rfila + ':' + rcolumna
        self.ws.unmerge_cells( rango )
        self.wb.save( nombre )


    def insertar_imagen( self , ruta  , celda):
        img = openpyxl.drawing.image.image( ruta )
        self.ws.add_image( img , celda )
        self.wb.save( nombre )


    def insertar_rangofila( self , rango , fila , columna ):
        for i , value in enumerate( rango ):
            self.ws.cell( column = columna + i , row = fila , value = value)
        self.wb.save( nombre )


    def insertar_rangocolumna(self , rango , fila , columna ):
        for i , value in enumerate( rango ):
            self.ws.cell( column = columna  , row = fila + i , value = value)
        self.wb.save( nombre )


    def cambiarcolorfila(self , fila , col , inicol , fincol ):
        for i in range( col ):
            self.ws.cell( row = fila , column = i + 1 ).fill = patternfill( start_color= inicol, end_color= fincol, fill_type = 'solid')
        self.wb.save(nombre)


    def cambiarcolorcolumna(self , fila , col , inicol , fincol ):
        for i in range( fila ):
            self.ws.cell( row = i + 1 , column = col ).fill = patternfill( start_color= inicol, end_color= fincol, fill_type = 'solid')
        self.wb.save(nombre)